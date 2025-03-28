import os
import sys
import logging
import csv
import json
from copy import copy
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from tqdm import tqdm
from utils import safe_convert_date
from config_manager import config_manager
from file_processor import FileProcessor, transform_month_column
from monetary_utils import standardize_monetary_columns, format_excel_monetary_columns
from time_utils import transform_times, excel_time_to_seconds, seconds_to_excel_time
from user_interface import (
    collect_user_inputs,
    verify_languages,
    print_header,
    select_processing_mode,
    display_batch_summary,
    choose_input_file,
    prompt_batch_settings,
    prompt_for_contract,
    prompt_for_estimate,
)


@dataclass
class ProcessingResult:
    """Tracks the result of processing a single file."""

    filename: str
    success: bool
    error_message: Optional[str] = None
    warnings: List[str] = field(default_factory=list)
    metrics: Dict = field(default_factory=dict)
    output_file: Optional[str] = None


class ProcessingError(Exception):
    """Custom exception for processing-related errors."""

    pass


class EtereBridge:
    """Enhanced file processor with error recovery and progress tracking."""

    def __init__(self):
        """Initialize the EtereBridge processor."""
        self.config = config_manager.get_config()
        self.log_file = config_manager.setup_logging()
        self.results: List[ProcessingResult] = []

        # Initialize FileProcessor
        self.file_processor = FileProcessor(self.config)

    def list_files(self) -> List[str]:
        """List all available files in the input directory."""
        files = [
            f for f in os.listdir(self.config.paths.input_dir) if f.endswith(".csv")
        ]
        if not files:
            print(
                "\n❌ No CSV files found in the input directory:",
                self.config.paths.input_dir,
            )
            print("Please add your CSV files to this directory and try again.")
            sys.exit(1)
        return files

    def get_worldlink_defaults(self) -> Dict:
        return {
            "billing_type": "Broadcast",
            "revenue_type": "Direct Response Sales",
            "agency_flag": "Agency",
            "sales_person": "House",
            "agency_fee": 0.15,
            "type": "COM",
            "affidavit": "Y",
            "is_worldlink": True,
            "estimate": "",  # default empty string
            "contract": "DEFAULT",  # or an empty string if appropriate
        }

    def extract_header_values(self, file_path: str) -> Tuple[str, str]:
        """
        Extract values for Bill Code generation from the new CSV format.
        
        From the example showing:
        > [Textbox180,COD_CONTRATTO,COD_CONTRATTO2,Textbox172,Textbox181,Textbox171,Textbox182,Detail]
        > Rod: Placcow Malin,RPM TVC 10596 SF,3/18/2025,Thunder Valley Casino Est 10596 SFO,"222 S. Morgan St, Ste 100",Thunder Valley Casino,Chicago,
        
        We need:
        - First part: Value from the first column of line 2 (client/agency name)
        - Second part: Value from the 6th column of line 2 (site/venue name)
        """
        try:
            logging.info(f"Extracting header values from: {file_path}")
            
            # Read the file line by line instead of reading the whole file at once
            with open(file_path, 'r') as f:
                # Skip the first line (column headers)
                f.readline()
                
                # Read the second line which contains our data
                second_line = f.readline().strip()
                
                if not second_line:
                    logging.error("Could not find data line in file")
                    return "", ""
                    
                logging.info(f"Processing line: {second_line}")
                
                # Use csv module to properly handle quoted fields
                import csv
                reader = csv.reader([second_line])
                parts = next(reader)
                
                # Extract first part (client/agency) from first column
                first_part = parts[0].strip() if len(parts) > 0 else ""
                
                # Extract second part (venue/site) from 6th column
                second_part = parts[5].strip() if len(parts) > 5 else ""
                
                # If we have no second part, try to extract it from another column
                # without assuming it contains "Casino"
                if not second_part and len(parts) > 3:
                    # Try column 4 which might contain a venue name
                    potential_venue = parts[3].strip() if len(parts) > 3 else ""
                    if potential_venue and "Est" not in potential_venue:  # Avoid the "Est" column
                        second_part = potential_venue
                
                logging.info(f"Extracted first part: '{first_part}', second part: '{second_part}'")
                return str(first_part), str(second_part)
                
        except Exception as e:
            logging.error(f"Error in extract_header_values: {e}")
            logging.exception(e)
            return "", ""

    def apply_user_inputs(
        self,
        df: pd.DataFrame,
        billing_type: str,
        revenue_type: str,
        agency_flag: str,
        sales_person: str,
        agency_fee: Optional[float],
        language: Dict,
        affidavit: str,
        estimate: str,
        contract: str,
        is_worldlink: bool = False,
    ) -> pd.DataFrame:
        """
        Apply user input values to the DataFrame.

        Adds columns for billing type, revenue type, agency flag, sales person,
        language, affidavit, and estimate, then handles WorldLink-specific
        processing and broker fees. Also computes the Type column automatically:
        - If the Gross Rate is blank or zero, Type is 'BNS'
        - Otherwise, Type is 'COM'
        Ensures all required columns exist and orders them according to configuration.
        """
        try:
            logging.info("Applying user inputs to DataFrame...")

            # Add user input columns
            df["Billing Type"] = billing_type
            df["Revenue Type"] = revenue_type
            df["Agency?"] = agency_flag
            df["Sales Person"] = sales_person
            df["Lang."] = df.index.map(language)
            df["Affidavit?"] = affidavit

            # Add Estimate and Contract columns from user input
            df["Estimate"] = estimate
            df["Contract"] = contract  # <-- New: write contract number

            # Compute Type automatically from Gross Rate on a per-row basis.
            def compute_type(row):
                try:
                    # Use the numeric value directly (no string conversion/parsing needed)
                    value = row.get("Gross Rate", 0)
                    # Ensure it's treated as a number even if it came from elsewhere
                    if not isinstance(value, (int, float)):
                        value = float(str(value).replace("$", "").replace(",", ""))

                    if value == 0:
                        return "BNS"
                    else:
                        return "COM"
                except Exception as e:
                    logging.warning(f"Error computing type for row: {e}")
                    return "BNS"

            df["Type"] = df.apply(compute_type, axis=1)

            # Handle WorldLink-specific processing
            if is_worldlink:
                logging.info("Processing WorldLink order specific requirements...")
                if "Market" in df.columns:
                    logging.info("Copying Market data to Makegood column")
                    if "Make Good" not in df.columns:
                        df["Make Good"] = None
                    df["Make Good"] = df["Market"]
                    logging.info("Successfully copied Market data to Make Good")
                else:
                    logging.warning(
                        "Market column not found in WorldLink order - cannot copy to Make Good"
                    )

            # Handle agency fees
            if agency_flag == "Agency" and agency_fee is not None:
                df["Broker Fees"] = None
            else:
                df["Broker Fees"] = None

            # Ensure all required columns exist
            logging.info("Ensuring all required columns exist...")
            for col in self.config.final_columns:
                if col not in df.columns:
                    logging.info(f"Adding missing column: {col}")
                    df[col] = None

            # Reorder columns according to configuration
            logging.info("Reordering columns according to configuration...")
            try:
                df = df[self.config.final_columns]
            except KeyError as e:
                missing_cols = [
                    col for col in self.config.final_columns if col not in df.columns
                ]
                logging.error(f"Missing columns: {missing_cols}")
                raise KeyError(f"Missing required columns: {missing_cols}")

            logging.info("Successfully applied user inputs!")
            return df

        except Exception as e:
            logging.error(f"Error applying user inputs: {str(e)}")
            raise

    def save_to_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        agency_fee: Optional[float] = 0.15
    ):
        """
        Write the final DataFrame 'df' to an Excel file, preserving formulas,
        template formatting, and special handling for time/date columns.

        Args:
            df (pd.DataFrame): The final, transformed data to be written to Excel.
            output_path (str): The path where the completed Excel file is saved.
            agency_fee (float, optional): Agency fee used to build broker-fee formulas if Agency? = Agency.
        """
        try:
            # 1) Load the template workbook
            template_path = self.config.paths.template_path
            logging.info(f"Loading template from: {template_path}")
            workbook = load_workbook(template_path, data_only=False)
            sheet = workbook.active

            # Gather final column order from config
            columns = self.config.final_columns

            # 2) Extract formulas and formatting from template row 2
            template_formulas = {}
            template_formatting = {}
            for col in range(1, len(columns) + 1):
                cell = sheet.cell(row=2, column=col)
                if cell.value and str(cell.value).startswith("="):
                    template_formulas[col] = cell.value
                template_formatting[col] = {
                    "style": cell.style,
                    "number_format": cell.number_format,
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "font": copy(cell.font),
                    "alignment": copy(cell.alignment),
                }

            # 3) Write headers in row 1
            for col_num, column_title in enumerate(columns, start=1):
                sheet.cell(row=1, column=col_num, value=column_title)

            # Identify columns for special formatting
            gross_col_index = None
            spot_value_col_index = None
            station_net_col_index = None
            broker_col_index = None
            gross_col_letter = None
            broker_col_letter = None
            air_date_idx = None
            air_date_letter = None
            agency_flag_idx = None
            agency_column_data = None

            # Define all monetary columns that need consistent currency formatting
            monetary_columns = ["Gross Rate", "Spot Value", "Station Net", "Broker Fees"]
            
            # Find column indices
            if "Gross Rate" in columns:
                gross_col_index = columns.index("Gross Rate") + 1
                gross_col_letter = get_column_letter(gross_col_index)
            if "Spot Value" in columns:
                spot_value_col_index = columns.index("Spot Value") + 1
            if "Station Net" in columns:
                station_net_col_index = columns.index("Station Net") + 1
            if "Broker Fees" in columns:
                broker_col_index = columns.index("Broker Fees") + 1
                broker_col_letter = get_column_letter(broker_col_index)
            if "Air Date" in columns:
                air_date_idx = columns.index("Air Date") + 1
                air_date_letter = get_column_letter(air_date_idx)
            if "Agency?" in columns:
                agency_flag_idx = columns.index("Agency?")
                agency_column_data = df[columns[agency_flag_idx]]

            # 4) Write data starting at row 2
            for row_num, row_data in enumerate(df.values, start=2):
                for col_num, cell_value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    col_name = columns[col_num - 1]

                    # A) Convert Time In/Time Out to numeric time
                    if col_name in ("Time In", "Time Out"):
                        time_serial = self._parse_time_24h(cell_value)
                        if time_serial is not None:
                            cell.value = time_serial
                            cell.number_format = "[h]:mm:ss"
                        else:
                            cell.value = cell_value

                    # B) End Date handling - always use the same date as Air Date
                    # and always apply the date formatting
                    elif col_name == "End Date":
                        if air_date_letter:
                            # Link to Air Date value and apply date formatting
                            cell.value = f"={air_date_letter}{row_num}"
                            cell.number_format = "m/d/yy"
                        else:
                            # If we can't link to Air Date, use the value directly 
                            # but still format it as a date
                            try:
                                dt = safe_convert_date(cell_value)
                                if dt is not None:
                                    cell.value = dt
                                    cell.number_format = "m/d/yy"
                                else:
                                    cell.value = cell_value
                            except Exception as e:
                                logging.warning(
                                    f"Error setting End Date at row {row_num}: {e}"
                                )
                                cell.value = cell_value

                    # C) Check if there's a template formula for this column
                    elif (
                        col_num in template_formulas
                        and col_name not in ("Time In", "Time Out", "Length", "End Date", "Broker Fees")
                    ):
                        formula = template_formulas[col_num]
                        cell.value = formula.replace("2", str(row_num))

                    # D) Inject Broker Fees formula if Agency? == "Agency"
                    elif col_name == "Broker Fees" and agency_fee is not None:
                        if agency_column_data is not None and gross_col_letter:
                            agency_flag_val = agency_column_data.iloc[row_num - 2]
                            if agency_flag_val == "Agency":
                                cell.value = f"={gross_col_letter}{row_num}*{agency_fee}"
                            else:
                                cell.value = None

                    # E) Length conversion to fraction-of-day
                    elif col_name == "Length":
                        try:
                            if pd.notna(cell_value):
                                length_in_seconds = float(cell_value)
                                time_fraction = length_in_seconds / 86400
                            else:
                                time_fraction = 0
                            cell.value = time_fraction
                            cell.number_format = "[h]:mm:ss"
                            cell.alignment = Alignment(horizontal="center")
                        except Exception as e:
                            logging.warning(
                                f"Error converting Length at row {row_num}: {e}. Storing raw value."
                            )
                            cell.value = cell_value
                    
                    # F) For monetary columns, just set the value as is
                    # (formatting will be handled by format_excel_monetary_columns later)
                    elif col_name in monetary_columns:
                        # We assume the df already has clean numeric values from standardize_monetary_columns
                        cell.value = cell_value if pd.notna(cell_value) else 0
                    
                    else:
                        cell.value = cell_value

                    # G) Apply template formatting
                    if col_num in template_formatting:
                        fmt = template_formatting[col_num]
                        cell.fill = fmt["fill"]
                        cell.border = fmt["border"]
                        cell.font = fmt["font"]
                        cell.alignment = fmt["alignment"]
                        if col_name not in ("Time In", "Time Out", "Length", "End Date") + tuple(monetary_columns):
                            cell.style = fmt["style"]
                            cell.number_format = fmt["number_format"]

            # Apply currency formatting to all monetary columns
            format_excel_monetary_columns(sheet, df, monetary_columns)

            # 5) Format Air Date as m/d/yy if present
            if "Air Date" in columns:
                air_date_col = columns.index("Air Date") + 1
                for row_num in range(2, len(df) + 2):
                    cell = sheet.cell(row=row_num, column=air_date_col)
                    if cell.value:
                        dt = safe_convert_date(cell.value)
                        if dt is not None:
                            cell.value = dt
                            cell.number_format = "m/d/yy"
                        else:
                            logging.warning(
                                f"Error formatting Air Date row {row_num}: value '{cell.value}' not parseable"
                            )

            # 6) Format Month if present
            if "Month" in columns:
                month_col = columns.index("Month") + 1
                for row_num in range(2, len(df) + 2):
                    cell = sheet.cell(row=row_num, column=month_col)
                    month_val = df["Month"].iloc[row_num - 2]
                    if pd.notna(month_val):
                        cell.value = month_val
                        cell.number_format = "mmm-yy"
                    else:
                        cell.value = None

            # 7) Set the Priority column to 4 if it exists
            if "Priority" in columns:
                priority_col = columns.index("Priority") + 1
                for row_num in range(2, len(df) + 2):
                    sheet.cell(row=row_num, column=priority_col, value=4)

            # 8) Remove extra template rows
            if sheet.max_row > len(df) + 1:
                sheet.delete_rows(len(df) + 2, sheet.max_row - (len(df) + 1))

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            workbook.save(output_path)
            logging.info("Excel file saved successfully with in-cell formulas and original template formatting.")

        except Exception as e:
            logging.error(f"Error saving to Excel: {str(e)}")
            raise

    def _parse_time_24h(self, time_str: str) -> Optional[float]:
        """
        Converts 'time_str' (24-hour or 12-hour) into an Excel time serial (a float).
        The result is (total seconds)/86400. Returns None if parsing fails.
        """
        if not time_str:
            return None
        try:
            dt = pd.to_datetime(time_str, format="%H:%M:%S", errors="raise")
        except ValueError:
            try:
                dt = pd.to_datetime(time_str, format="%I:%M:%S %p", errors="raise")
            except ValueError:
                return None
        total_seconds = dt.hour * 3600 + dt.minute * 60 + dt.second
        return total_seconds / 86400.0

    def generate_processing_summary(
        self, df: pd.DataFrame, input_file: str, output_file: str, user_inputs: Dict
    ) -> Dict:
        try:
            df["Air Date"] = df["Air Date"].apply(safe_convert_date)

            # Work with numeric Gross Rate values directly instead of string parsing
            gross_values = df["Gross Rate"]

            # If values are still strings with $ (for robustness), convert them
            if pd.api.types.is_string_dtype(gross_values):
                gross_values = pd.to_numeric(
                    gross_values.str.replace("$", "").str.replace(",", ""),
                    errors="coerce",
                ).fillna(0)

            df["Day_of_Week"] = df["Air Date"].dt.day_name()
            spots_by_day = df["Day_of_Week"].value_counts().to_dict()

            summary = {
                "processing_info": {
                    "timestamp": datetime.now().isoformat(),
                    "input_file": input_file,
                    "output_file": output_file,
                    "user_inputs": user_inputs,
                },
                "overall_metrics": {
                    "total_spots": len(df),
                    "total_gross_value": float(gross_values.sum()),
                    "average_spot_value": float(gross_values.mean()),
                    "unique_programs": len(df["Program"].unique()),
                },
                "date_range": {
                    "earliest": df["Air Date"].min().isoformat(),
                    "latest": df["Air Date"].max().isoformat(),
                    "total_days": (df["Air Date"].max() - df["Air Date"].min()).days
                    + 1,
                },
                "breakdowns": {
                    "markets": df["Market"].value_counts().to_dict(),
                    "media_types": df["Media"].value_counts().to_dict(),
                    "spots_by_day": spots_by_day,
                    "programs": df["Program"].value_counts().to_dict(),
                },
            }

            logging.info(f"Generated summary for {input_file}")
            return summary

        except Exception as e:
            logging.error(f"Error generating summary: {e}")
            raise

    def process_file(
        self, file_path: str, user_inputs: Optional[Dict] = None
    ) -> ProcessingResult:
        filename = os.path.basename(file_path)
        logging.info(f"###### Starting processing of {filename} ######")

        try:
            logging.info("Extracting header values...")
            text_box_180, text_box_171 = self.extract_header_values(file_path)
            logging.info("Loading and cleaning data...")
            df = self.file_processor.load_and_clean_data(file_path)
            logging.info("Detecting languages in data...")
            detected_counts, row_languages = self.file_processor.detect_languages(df)
            logging.info(f"Detected language counts: {detected_counts}")
            logging.info("Applying transformations...")
            df = self.file_processor.apply_transformations(
                df, text_box_180, text_box_171
            )

            # Add standardization of monetary columns
            logging.info("Standardizing monetary columns...")
            df = standardize_monetary_columns(df)
        
            # Transform time columns
            logging.info("Standardizing time formats...")
            df = transform_times(df)

            if user_inputs is None:
                logging.info("Collecting user inputs...")
                user_inputs = collect_user_inputs(self.config)
                user_inputs["is_worldlink"] = False

            language_dict = row_languages.to_dict() if not row_languages.empty else {}
            user_inputs["language"] = language_dict

            logging.info("Verifying languages...")
            primary_language = verify_languages(df, (detected_counts, row_languages))
            if isinstance(primary_language, pd.Series):
                primary_language = primary_language.to_dict()
            user_inputs["language"] = primary_language

            logging.info("Applying user inputs...")
            df = self.apply_user_inputs(
                df,
                billing_type=user_inputs["billing_type"],
                revenue_type=user_inputs["revenue_type"],
                agency_flag=user_inputs["agency_flag"],
                sales_person=user_inputs["sales_person"],
                agency_fee=user_inputs["agency_fee"],
                language=user_inputs["language"],
                affidavit=user_inputs["affidavit"],
                estimate=user_inputs["estimate"],
                contract=user_inputs["contract"],
                is_worldlink=user_inputs.get("is_worldlink", False),
            )
            df = transform_month_column(df)

            logging.info("Saving output file...")
            output_filename = f"processed_{os.path.splitext(filename)[0]}.xlsx"
            output_path = os.path.join(self.config.paths.output_dir, output_filename)
            self.save_to_excel(df, output_path, user_inputs.get("agency_fee"))

            logging.info("Generating processing summary...")
            summary = self.generate_processing_summary(
                df, file_path, output_path, user_inputs
            )
            language_distribution = (
                row_languages.value_counts().to_dict()
                if not row_languages.empty
                else {}
            )
            summary["language_info"] = {
                "detected_languages": detected_counts,
                "language_distribution": language_distribution,
            }

            if user_inputs.get("is_worldlink", False):
                summary["processing_info"]["worldlink_order"] = True
                if "Market" in df.columns:
                    summary["processing_info"]["market_to_makegood"] = "copied"
                else:
                    summary["processing_info"][
                        "market_to_makegood"
                    ] = "failed - Market column not found"

            return ProcessingResult(
                filename=filename,
                success=True,
                output_file=output_path,
                metrics=summary,
            )

        except FileNotFoundError as e:
            error_msg = f"File not found: {filename}"
            logging.error(error_msg)
            return ProcessingResult(
                filename=filename, success=False, error_message=error_msg
            )
        except pd.errors.EmptyDataError as e:
            error_msg = f"File is empty: {filename}"
            logging.error(error_msg)
            return ProcessingResult(
                filename=filename, success=False, error_message=error_msg
            )
        except ProcessingError as e:
            error_msg = f"Processing error in {filename}: {str(e)}"
            logging.error(error_msg)
            return ProcessingResult(
                filename=filename, success=False, error_message=error_msg
            )
        except Exception as e:
            error_msg = f"Unexpected error processing {filename}: {str(e)}"
            logging.error(error_msg, exc_info=True)
            return ProcessingResult(
                filename=filename, success=False, error_message=error_msg
            )

    def process_batch(self, files: List[str], show_progress: bool = True) -> dict:
        successful = []
        failed = []

        batch_settings = prompt_batch_settings(self.config)
        is_worldlink = batch_settings.get("is_worldlink", False)

        # For non-worldlink, use shared inputs if available.
        base_user_inputs = None
        if not is_worldlink:
            base_user_inputs = batch_settings.get("inputs") or None

        files_iter = tqdm(files, desc="Processing files") if show_progress else files

        for file_path in files_iter:
            try:
                print(f"\nProcessing file: {os.path.basename(file_path)}")

                # Build file-specific user inputs.
                if is_worldlink:
                    file_inputs = self.get_worldlink_defaults()
                    # Per-file prompts for contract, estimate
                    file_inputs["contract"] = prompt_for_contract()
                    file_inputs["estimate"] = prompt_for_estimate()
                else:
                    # If the batch shares user inputs, clone them; else prompt
                    file_inputs = (
                        base_user_inputs.copy()
                        if base_user_inputs
                        else collect_user_inputs(self.config)
                    )

                # Now let process_file() handle all detection & transformations
                result = self.process_file(file_path, file_inputs)

                # Sort the result
                if result.success:
                    successful.append(result)
                else:
                    failed.append(result)

                self._save_interim_results(successful, failed)

            except Exception as e:
                logging.error(f"Error processing {file_path}: {str(e)}")
                failed.append(
                    ProcessingResult(
                        filename=os.path.basename(file_path),
                        success=False,
                        error_message=str(e),
                    )
                )

        display_batch_summary(successful, failed, self.log_file)
        return {"successful": successful, "failed": failed}

    def _save_interim_results(
        self, successful: List[ProcessingResult], failed: List[ProcessingResult]
    ):
        interim_file = Path(self.config.paths.output_dir) / "interim_results.json"
        results = {
            "timestamp": datetime.now().isoformat(),
            "successful": [],
            "failed": [],
        }
        for result in successful:
            result_dict = vars(result)
            if (
                "metrics" in result_dict
                and "language_distribution" in result_dict["metrics"]
            ):
                result_dict["metrics"]["language_distribution"] = result_dict[
                    "metrics"
                ]["language_distribution"].to_dict()
            results["successful"].append(result_dict)
        for result in failed:
            results["failed"].append(vars(result))
        with open(interim_file, "w") as f:
            json.dump(results, f, indent=2)

    def main(self):
        print_header(self.log_file)

        try:
            files = self.list_files()
            if not files:
                print("No files found to process. Please add files and try again.")
                return

            choice = select_processing_mode()

            if choice == "A":
                print("\n🔄 Processing all files automatically...")
                file_paths = [
                    os.path.join(self.config.paths.input_dir, f) for f in files
                ]
                results = self.process_batch(file_paths)
            elif choice == "S":
                while True:
                    file_path = choose_input_file(files, self.config.paths.input_dir)
                    if file_path:
                        results = self.process_batch([file_path], show_progress=False)
                    print("\n" + "-" * 80)
                    cont = (
                        input("\nWould you like to process another file? (Y/N): ")
                        .strip()
                        .lower()
                    )
                    if cont != "y":
                        print("\n✅ Processing complete! Thank you for using the tool.")
                        break

        except KeyboardInterrupt:
            print("\n\nProgram interrupted by user. Saving interim results...")
            self._save_interim_results(self.results, [])
            print("Interim results saved. Exiting...")
            sys.exit(0)
        except Exception as e:
            logging.error(f"Unexpected error: {str(e)}")
            print(
                f"\n❌ An unexpected error occurred. Please check the log file: {self.log_file}"
            )
            sys.exit(1)


if __name__ == "__main__":
    processor = EtereBridge()
    processor.main()
