import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import math
import sys
import csv
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from tqdm import tqdm
from config_manager import config_manager

@dataclass
class ProcessingResult:
    """Tracks the result of processing a single file."""
    filename: str
    success: bool
    error_message: Optional[str] = None
    warnings: List[str] = field(default_factory=list)
    metrics: Dict = field(default_factory=dict)
    output_file: Optional[str] = None

class ProcessingError(Exception):
    """Custom exception for processing-related errors."""
    pass

class EtereBridge:
    """Enhanced file processor with error recovery and progress tracking."""
    
    def __init__(self):
        """Initialize the EtereBridge processor."""
        self.config = config_manager.get_config()
        self._setup_logging()
        self.results: List[ProcessingResult] = []
        
    def _setup_logging(self):
        """Configure logging with both file and console output."""
        log_dir = Path(self.config.paths.output_dir) / 'logs'
        log_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate timestamp-based log filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f'processing_{timestamp}.log'
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        
        self.log_file = log_file
        logging.info(f"Starting new processing session. Log file: {log_file}")

    def print_header(self):
        """Display a welcome header with basic instructions."""
        header = """
╔════════════════════════════════════════════════════════════════════════════╗
║                        Excel File Processing Tool                           ║
╚════════════════════════════════════════════════════════════════════════════╝

This tool helps you process and transform Excel files according to specified formats.
Follow the prompts below to begin processing your files.

Version: 2.0
Log File: {log_file}
        """.format(log_file=self.log_file)
        print(header)

    def list_files(self) -> List[str]:
        """List all available files in the input directory."""
        files = [f for f in os.listdir(self.config.paths.input_dir) 
                if f.endswith('.csv')]
        if not files:
            print("\n❌ No CSV files found in the input directory:", 
                  self.config.paths.input_dir)
            print("Please add your CSV files to this directory and try again.")
            sys.exit(1)
        return files

    def select_processing_mode(self) -> str:
        """Ask the user whether to process all files or select one at a time."""
        print("\n" + "-"*80)
        print("Processing Mode Selection".center(80))
        print("-"*80)
        print("\nChoose how you want to process your files:")
        print("  [A] Process all files automatically")
        print("  [S] Select and process files one at a time")
        
        while True:
            choice = input("\nYour choice (A/S): ").strip().upper()
            if choice in ['A', 'S']:
                return choice
            print("❌ Invalid choice. Please enter 'A' for all files or 'S' to select files.")

    def select_input_file(self, files: List[str]) -> Optional[str]:
        """Prompt the user to select a file from the input directory."""
        print("\n" + "-"*80)
        print("File Selection".center(80))
        print("-"*80)
        print("\nAvailable files for processing:")
        
        # Create two columns if there are many files
        mid_point = (len(files) + 1) // 2
        for i, filename in enumerate(files, 1):
            line = f"  [{i:2d}] {filename}"
            if i <= mid_point and i + mid_point <= len(files):
                second_file = files[i + mid_point - 1]
                second_item = f"  [{i + mid_point:2d}] {second_file}"
                print(f"{line:<40} {second_item}")
            else:
                print(line)
        
        while True:
            try:
                choice = input("\nEnter the number of the file you want to process (or 'q' to quit): ").strip()
                if choice.lower() == 'q':
                    print("\nExiting program...")
                    sys.exit(0)
                
                choice = int(choice)
                if 1 <= choice <= len(files):
                    selected_file = files[choice - 1]
                    print(f"\n✅ Selected: {selected_file}")
                    return os.path.join(self.config.paths.input_dir, selected_file)
                else:
                    print(f"❌ Please enter a number between 1 and {len(files)}")
            except ValueError:
                print("❌ Please enter a valid number or 'q' to quit")

    def extract_header_values(self, file_path: str) -> Tuple[str, str]:
        """Extract header values from first section of CSV."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()[:2]
                
                header_row = [x.strip() for x in lines[0].split(',')]
                value_row = next(csv.reader([lines[1]]))
                
                header_dict = dict(zip(header_row, value_row))
                
                text_box_180 = header_dict.get('Textbox180', '').strip()
                text_box_171 = header_dict.get('Textbox171', '').strip()
                
                logging.info(f"Header values found - TextBox180: '{text_box_180}', TextBox171: '{text_box_171}'")
                
                return text_box_180, text_box_171
                
        except Exception as e:
            logging.error(f"Error reading header: {e}")
            return '', ''

    def clean_numeric(self, value):
        """Clean numeric strings by removing commas and decimal points."""
        if isinstance(value, str):
            return value.replace(',', '').split('.')[0]
        return value

    def round_to_nearest_30_seconds(self, seconds):
        """Round the given number of seconds to the nearest 30-second increment."""
        try:
            if pd.isna(seconds) or not str(seconds).strip():
                return 0
            return round(float(seconds) / 15) * 15
        except (ValueError, TypeError) as e:
            logging.warning(f"Error rounding seconds '{seconds}': {e}")
            return 0

    def load_and_clean_data(self, file_path: str) -> Optional[pd.DataFrame]:
        """Load data from the selected input file and perform initial transformations."""
        try:
            logging.info(f"Loading data from {file_path}")
            
            # Read the main data
            df = pd.read_csv(file_path, skiprows=3)
            original_count = len(df)
            
            # Drop empty rows
            df = df.dropna(how='all')
            if len(df) < original_count:
                logging.warning(f"Dropped {original_count - len(df)} empty rows")
            
            # Filter required columns
            required_columns = ['id_contrattirighe', 'timerange2', 'dateschedule']
            before_required = len(df)
            df = df[df[required_columns].notna().all(axis=1)]
            if len(df) < before_required:
                logging.warning(f"Dropped {before_required - len(df)} rows with missing required values")
            
            # Filter out Textbox rows
            df = df[~df['IMPORTO2'].astype(str).str.contains('Textbox', na=False)]
            df = df[df.columns[~df.columns.str.contains('Textbox97|tot|Textbox61|Textbox53')]]
            
            # Clean numeric fields
            df['id_contrattirighe'] = df['id_contrattirighe'].apply(self.clean_numeric)
            if 'Textbox14' in df.columns:
                df['Textbox14'] = df['Textbox14'].apply(self.clean_numeric)
            
            # Rename columns
            column_mapping = {
                'id_contrattirighe': 'Line',
                'Textbox14': '#',
                'duration3': 'Length',
                'IMPORTO2': 'Gross Rate',
                'nome2': 'Market',
                'dateschedule': 'Air Date',
                'airtimep': 'Program',
                'bookingcode2': 'Media'
            }
            
            logging.info(f"Available columns before renaming: {df.columns.tolist()}")
            rename_dict = {k: v for k, v in column_mapping.items() if k in df.columns}
            df = df.rename(columns=rename_dict)
            logging.info(f"Columns after renaming: {df.columns.tolist()}")
            
            # Split timerange2
            if 'timerange2' in df.columns:
                df[['Time In', 'Time Out']] = df['timerange2'].str.split('-', expand=True)
            
            # Validate required columns after renaming
            df = df[df['Line'].notna()]
            
            if df.empty:
                raise ProcessingError("No valid data rows found after cleaning")
                
            return df
            
        except Exception as e:
            logging.error(f"Error in load_and_clean_data: {str(e)}")
            raise

    def generate_billcode(self, text_box_180: str, text_box_171: str) -> str:
        """Combine Textbox180 and Textbox171 for billcode."""
        if text_box_180 and text_box_171:
            return f"{text_box_180}:{text_box_171}"
        elif text_box_171:
            return text_box_171
        elif text_box_180:
            return text_box_180
        return ''

    def apply_transformations(self, df: pd.DataFrame, text_box_180: str, 
                            text_box_171: str) -> pd.DataFrame:
        """Apply transformations including billcode."""
        try:
            # Set billcode
            billcode = self.generate_billcode(text_box_180, text_box_171)
            df['Bill Code'] = billcode
            
            # Check Market column
            if 'Market' not in df.columns:
                logging.error("Market column not found in DataFrame")
                logging.info(f"Available columns: {df.columns.tolist()}")
                raise KeyError("Market column not found in DataFrame")
            
            # Apply market replacements
            logging.info(f"Applying market replacements: {self.config.market_replacements}")
            df['Market'] = df['Market'].replace(self.config.market_replacements)
            
            # Transform Gross Rate
            if 'Gross Rate' in df.columns:
                df['Gross Rate'] = df['Gross Rate'].astype(str).str.replace('$', '').str.replace(',', '')
                df['Gross Rate'] = pd.to_numeric(df['Gross Rate'], errors='coerce').fillna(0).map("${:,.2f}".format)
            
            # Transform Length
            if 'Length' in df.columns:
                df['Length'] = df['Length'].apply(self.round_to_nearest_30_seconds)
                df['Length'] = pd.to_timedelta(df['Length'], unit='s').apply(lambda x: str(x).split()[-1].zfill(8))
            
            # Transform Line and #
            if 'Line' in df.columns:
                df['Line'] = pd.to_numeric(df['Line'], errors='coerce').fillna(0).astype(int)
            
            if '#' in df.columns:
                df['#'] = pd.to_numeric(df['#'], errors='coerce').fillna(0).astype(int)
            
            return df
            
        except Exception as e:
            logging.error(f"Error in transformations: {str(e)}")
            raise

    def prompt_for_user_inputs(self) -> Dict:
        """Prompt the user for processing parameters."""
        print("\n" + "-"*80)
        print("Additional Information Needed".center(80))
        print("-"*80)
        
        # Get Sales Person
        sales_people = self.config.sales_people
        print("\n1. Sales Person:")
        for idx, person in enumerate(sales_people, 1):
            print(f"   [{idx}] {person}")
            
        while True:
            try:
                choice = int(input("\nSelect sales person (enter number): "))
                if 1 <= choice <= len(sales_people):
                    sales_person = sales_people[choice-1]
                    break
                print(f"❌ Please enter a number between 1 and {len(sales_people)}")
            except ValueError:
                print("❌ Please enter a valid number")
        
        # Billing Type
        print("\n2. Billing Type:")
        print("   [C] Calendar")
        print("   [B] Broadcast")
        while True:
            billing_input = input("\nSelect billing type (C/B): ").strip().upper()
            if billing_input in ['C', 'B']:
                billing_type = "Calendar" if billing_input == 'C' else "Broadcast"
                break
            print("❌ Please enter 'C' for Calendar or 'B' for Broadcast")
        
        # Revenue Type
        print("\n3. Revenue Type:")
        print("   [B] Branded Content")
        print("   [D] Direct Response")
        print("   [I] Internal Ad Sales")
        print("   [P] Paid Programming")
        while True:
            revenue_input = input("\nSelect revenue type (B/D/I/P): ").strip().upper()
            if revenue_input in ['B', 'D', 'I', 'P']:
                revenue_types = {
                    'B': "Branded Content",
                    'D': "Direct Response",
                    'I': "Internal Ad Sales",
                    'P': "Paid Programming"
                }
                revenue_type = revenue_types[revenue_input]
                break
            print("❌ Please enter 'B', 'D', 'I', or 'P'")
        
        # Agency Type and Fee
        print("\n4. Order Type:")
        print("   [A] Agency")
        print("   [N] Non-Agency")
        print("   [T] Trade")
        
        agency_fee = None
        while True:
            agency_input = input("\nSelect order type (A/N/T): ").strip().upper()
            if agency_input in ['A', 'N', 'T']:
                agency_types = {
                    'A': "Agency",
                    'N': "Non-Agency",
                    'T': "Trade"
                }
                agency_flag = agency_types[agency_input]
                
                if agency_input == 'A':
                    print("\n5. Agency Fee Type:")
                    print("   [S] Standard (15%)")
                    print("   [C] Custom")
                    while True:
                        fee_type = input("\nSelect fee type (S/C): ").strip().upper()
                        if fee_type == 'S':
                            agency_fee = 0.15
                            break
                        elif fee_type == 'C':
                            while True:
                                try:
                                    custom_fee = float(input("\nEnter custom fee percentage (without % symbol): "))
                                    if 0 <= custom_fee <= 100:
                                        agency_fee = custom_fee / 100
                                        break
                                    print("❌ Please enter a percentage between 0 and 100")
                                except ValueError:
                                    print("❌ Please enter a valid number")
                            break
                        print("❌ Please enter 'S' for Standard or 'C' for Custom")
                break
            print("❌ Please enter 'A' for Agency, 'N' for Non-Agency, or 'T' for Trade")
        
        print("\n✅ Information collected successfully!")
        return {
            "billing_type": billing_type,
            "revenue_type": revenue_type,
            "agency_flag": agency_flag,
            "sales_person": sales_person,
            "agency_fee": agency_fee
        }
    
    # Add this method to the EtereBridge class, near the other DataFrame manipulation methods
    def apply_user_inputs(self, df: pd.DataFrame, billing_type: str, revenue_type: str, 
                        agency_flag: str, sales_person: str, agency_fee: Optional[float]) -> pd.DataFrame:
        """Apply user input to the appropriate columns in the DataFrame."""
        try:
            logging.info("Applying user inputs to DataFrame...")
            
            # Add user input columns
            df['Billing Type'] = billing_type
            df['Revenue Type'] = revenue_type
            df['Agency?'] = agency_flag
            df['Sales Person'] = sales_person

            # Initialize Broker Fees column
            df['Broker Fees'] = None

            # Ensure all required columns exist
            logging.info("Ensuring all required columns exist...")
            for col in self.config.final_columns:
                if col not in df.columns:
                    logging.info(f"Adding missing column: {col}")
                    df[col] = None
            
            # Reorder columns according to config
            logging.info("Reordering columns according to configuration...")
            try:
                df = df[self.config.final_columns]
            except KeyError as e:
                missing_cols = [col for col in self.config.final_columns if col not in df.columns]
                logging.error(f"Missing columns: {missing_cols}")
                raise KeyError(f"Missing required columns: {missing_cols}")
            
            logging.info("Successfully applied user inputs!")
            return df
            
        except Exception as e:
            logging.error(f"Error applying user inputs: {str(e)}")
            raise
    
    # Add these methods to the EtereBridge class
    def save_output_file(self, df: pd.DataFrame, input_file: str, 
                        user_inputs: Dict) -> str:
        """Save the processed DataFrame to an Excel file."""
        try:
            # Define output file name
            filename_base = os.path.splitext(os.path.basename(input_file))[0]
            timestamp = datetime.now().strftime("%Y-%m-%d")
            output_file = os.path.join(
                self.config.paths.output_dir, 
                f"{filename_base}_Processed_{timestamp}.xlsx"
            )

            self.save_to_excel(
                df, 
                self.config.paths.template_path, 
                output_file, 
                user_inputs.get('agency_fee')
            )
            
            logging.info(f"Successfully saved output to: {output_file}")
            return output_file
            
        except Exception as e:
            logging.error(f"Error saving output file: {e}")
            raise

    def save_to_excel(self, df: pd.DataFrame, template_path: str, 
                    output_path: str, agency_fee: Optional[float] = 0.15):
        """Save DataFrame to Excel, preserving template and handling formulas."""
        try:
            logging.info(f"Loading template from: {template_path}")
            workbook = load_workbook(template_path)
            sheet = workbook.active
            
            # Get column indices from config
            columns = self.config.final_columns
            broker_fees_idx = columns.index('Broker Fees') + 1 if 'Broker Fees' in columns else None
            gross_rate_idx = columns.index('Gross Rate') + 1 if 'Gross Rate' in columns else None
            agency_idx = columns.index('Agency?') + 1 if 'Agency?' in columns else None
            
            # Write headers
            for col_num, column_title in enumerate(columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)

            # Write data with enhanced error handling
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    try:
                        column_name = columns[col_num - 1]
                        cell = sheet.cell(row=row_num, column=col_num)
                        
                        if (column_name == 'Broker Fees' and broker_fees_idx and 
                            agency_idx and gross_rate_idx):
                            agency_cell = sheet.cell(row=row_num, column=agency_idx)
                            if agency_cell.value == 'Agency' and agency_fee is not None:
                                from openpyxl.utils import get_column_letter
                                gross_rate_col = get_column_letter(gross_rate_idx)
                                cell.value = f'={gross_rate_col}{row_num}*{agency_fee}'
                                cell.number_format = '$#,##0.00'
                            else:
                                cell.value = None
                        
                        elif column_name in ['Gross Rate', 'Spot Value', 'Station Net']:
                            if cell_value and str(cell_value).strip():
                                clean_value = str(cell_value).replace('$', '').replace(',', '')
                                try:
                                    cell.value = float(clean_value)
                                    cell.number_format = '$#,##0.00'
                                except ValueError:
                                    logging.warning(f"Could not convert {cell_value} to float for {column_name}")
                                    cell.value = cell_value
                        
                        else:
                            cell.value = cell_value

                    except Exception as e:
                        logging.error(f"Error writing cell at row {row_num}, col {col_num}: {e}")
                        cell.value = cell_value  # Fall back to direct value assignment

            # Remove excess rows while preserving formulas
            last_data_row = len(df) + 1
            if sheet.max_row > last_data_row:
                sheet.delete_rows(last_data_row + 1, sheet.max_row - last_data_row)
            
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save workbook
            workbook.save(output_path)
            logging.info("Excel file saved successfully with formulas")
            
        except Exception as e:
            logging.error(f"Error saving to Excel: {str(e)}")
            raise

    def generate_processing_summary(self, df: pd.DataFrame, input_file: str,
                                output_file: str, user_inputs: Dict) -> Dict:
        """Generate comprehensive summary of the processing results."""
        try:
            # Convert date column to datetime
            df['Air Date'] = pd.to_datetime(df['Air Date'])
            
            # Convert Gross Rate to numeric for calculations
            gross_values = pd.to_numeric(
                df['Gross Rate'].str.replace('$', '').str.replace(',', ''),
                errors='coerce'
            ).fillna(0)
            
            # Calculate spots by day of week
            df['Day_of_Week'] = df['Air Date'].dt.day_name()
            spots_by_day = df['Day_of_Week'].value_counts().to_dict()
            
            summary = {
                "processing_info": {
                    "timestamp": datetime.now().isoformat(),
                    "input_file": input_file,
                    "output_file": output_file,
                    "user_inputs": user_inputs
                },
                "overall_metrics": {
                    "total_spots": len(df),
                    "total_gross_value": float(gross_values.sum()),
                    "average_spot_value": float(gross_values.mean()),
                    "unique_programs": len(df['Program'].unique()),
                },
                "date_range": {
                    "earliest": df['Air Date'].min().isoformat(),
                    "latest": df['Air Date'].max().isoformat(),
                    "total_days": (df['Air Date'].max() - df['Air Date'].min()).days + 1
                },
                "breakdowns": {
                    "markets": df['Market'].value_counts().to_dict(),
                    "media_types": df['Media'].value_counts().to_dict(),
                    "spots_by_day": spots_by_day,
                    "programs": df['Program'].value_counts().to_dict()
                }
            }
            
            logging.info(f"Generated summary for {input_file}")
            return summary
            
        except Exception as e:
            logging.error(f"Error generating summary: {e}")
            raise

    def process_file(self, file_path: str) -> ProcessingResult:
        """Process a single input file with enhanced error handling."""
        filename = os.path.basename(file_path)
        logging.info(f"###### Starting processing of {filename} ######")
        
        try:
            # Extract header values
            logging.info("Extracting header values...")
            text_box_180, text_box_171 = self.extract_header_values(file_path)
            
            # Load and clean data
            logging.info("Loading and cleaning data...")
            df = self.load_and_clean_data(file_path)
            
            # Apply transformations
            logging.info("Applying transformations...")
            df = self.apply_transformations(df, text_box_180, text_box_171)
            
            # Get user inputs
            logging.info("Collecting user inputs...")
            user_inputs = self.prompt_for_user_inputs()
            
            # Apply user inputs
            logging.info("Applying user inputs...")
            df = self.apply_user_inputs(df, **user_inputs)
            
            # Save output file
            logging.info("Saving output file...")
            output_file = self.save_output_file(df, file_path, user_inputs)
            
            # Generate and save summary
            logging.info("Generating processing summary...")
            summary = self.generate_processing_summary(df, file_path, output_file, user_inputs)
            
            return ProcessingResult(
                filename=filename,
                success=True,
                output_file=output_file,
                metrics=summary
            )
            
        except Exception as e:
            logging.error(f"Error processing {filename}: {str(e)}")
            return ProcessingResult(
                filename=filename,
                success=False,
                error_message=str(e)
            )

    def process_batch(self, files: List[str], show_progress: bool = True) -> Dict[str, List[ProcessingResult]]:
        """Process multiple files with progress tracking and error recovery."""
        successful = []
        failed = []
        
        files_iter = tqdm(files, desc="Processing files") if show_progress else files
        
        for file_path in files_iter:
            result = self.process_file(file_path)
            if result.success:
                successful.append(result)
            else:
                failed.append(result)
            
            # Save interim results
            self._save_interim_results(successful, failed)

        self._display_batch_summary(successful, failed)
        return {"successful": successful, "failed": failed}

    def _save_interim_results(self, successful: List[ProcessingResult], 
                            failed: List[ProcessingResult]):
        """Save interim results to protect against crashes."""
        interim_file = Path(self.config.paths.output_dir) / 'interim_results.json'
        
        results = {
            "timestamp": datetime.now().isoformat(),
            "successful": [vars(r) for r in successful],
            "failed": [vars(r) for r in failed]
        }
        
        with open(interim_file, 'w') as f:
            json.dump(results, f, indent=2)

    def _display_batch_summary(self, successful: List[ProcessingResult], 
                             failed: List[ProcessingResult]):
        """Display a user-friendly summary of batch processing results."""
        print("\n" + "="*80)
        print("Batch Processing Summary".center(80))
        print("="*80)
        
        # Success/failure counts
        total = len(successful) + len(failed)
        success_rate = (len(successful) / total * 100) if total > 0 else 0
        
        print(f"\nTotal files processed: {total}")
        print(f"Successfully processed: {len(successful)} ({success_rate:.1f}%)")
        print(f"Failed to process: {len(failed)}")
        
        # Display failures
        if failed:
            print("\nFailed Files:")
            for result in failed:
                print(f"❌ {result.filename}")
                print(f"   Error: {result.error_message}")
        
        # Display warnings
        if any(r.warnings for r in successful):
            print("\nWarnings:")
            for result in successful:
                if result.warnings:
                    print(f"⚠️ {result.filename}:")
                    for warning in result.warnings:
                        print(f"   - {warning}")

        # Display output locations
        if successful:
            print("\nProcessed Files:")
            for result in successful:
                print(f"✅ {result.filename} -> {result.output_file}")

        print(f"\nDetailed logs available at: {self.log_file}")

    def main(self):
        """Main function to control the flow of the program."""
        self.print_header()
        
        try:
            files = self.list_files()
            if not files:
                print("No files found to process. Please add files and try again.")
                return

            choice = self.select_processing_mode()

            if choice == 'A':
                print("\n🔄 Processing all files automatically...")
                file_paths = [os.path.join(self.config.paths.input_dir, f) for f in files]
                results = self.process_batch(file_paths)
                
            elif choice == 'S':
                while True:
                    file_path = self.select_input_file(files)
                    if file_path:
                        results = self.process_batch([file_path], show_progress=False)
                    
                    print("\n" + "-"*80)
                    cont = input("\nWould you like to process another file? (Y/N): ").strip().lower()
                    if cont != 'y':
                        print("\n✅ Processing complete! Thank you for using the tool.")
                        break

        except KeyboardInterrupt:
            print("\n\nProgram interrupted by user. Saving interim results...")
            self._save_interim_results(self.results, [])
            print("Interim results saved. Exiting...")
            sys.exit(0)
        except Exception as e:
            logging.error(f"Unexpected error: {str(e)}")
            print(f"\n❌ An unexpected error occurred. Please check the log file: {self.log_file}")
            sys.exit(1)

if __name__ == "__main__":
    processor = EtereBridge()
    processor.main()
