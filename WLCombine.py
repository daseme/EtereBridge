import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

def combine_excel_files_convert_formulas(folder_path, output_file='combined_output.xlsx'):
    """
    1. Opens each file and converts all formulas to values (paste special)
    2. Combines all files: first file gets all rows, subsequent files skip header
    3. Preserves formatting
    """
    
    # Get all Excel files in the folder
    excel_files = []
    for file in os.listdir(folder_path):
        if file.endswith(('.xlsx', '.xlsm')) and file != output_file:
            excel_files.append(os.path.join(folder_path, file))
    
    if not excel_files:
        print("No Excel files found in the folder.")
        return
    
    print(f"Found {len(excel_files)} Excel files")
    print("\nStep 1: Converting formulas to values in each file...")
    
    # Step 1: Convert formulas to values in each file
    for file_path in excel_files:
        try:
            print(f"  Converting formulas in: {os.path.basename(file_path)}")
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Iterate through all cells and convert formulas to values
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # 'f' means formula
                        # Replace formula with its calculated value
                        cell.value = cell.value
            
            # Save the file with formulas converted to values
            wb.save(file_path)
            wb.close()
            print(f"    ✓ Saved with values only")
            
        except Exception as e:
            print(f"  Error converting {file_path}: {e}")
    
    print("\nStep 2: Combining all files into output...")
    
    # Step 2: Create combined output file
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Combined Data"
    
    current_row = 1
    
    # Process each file
    for file_idx, file_path in enumerate(excel_files):
        try:
            print(f"  Processing: {os.path.basename(file_path)}")
            wb = load_workbook(file_path)
            ws = wb.active
            
            # First file: copy ALL rows (including header)
            # Subsequent files: skip header (start from row 2)
            start_row = 1 if file_idx == 0 else 2
            
            # Copy rows
            for row in ws.iter_rows(min_row=start_row):
                for col_idx, cell in enumerate(row, start=1):
                    new_cell = output_ws.cell(row=current_row, column=col_idx)
                    
                    # Copy value
                    new_cell.value = cell.value
                    
                    # Copy formatting
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.number_format = copy(cell.number_format)
                    
                    # Copy column width from first file
                    if file_idx == 0 and current_row == 1:
                        col_letter = get_column_letter(col_idx)
                        if col_letter in ws.column_dimensions:
                            output_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                
                current_row += 1
            
            wb.close()
            print(f"    ✓ Added {ws.max_row - start_row + 1} rows")
            
        except Exception as e:
            print(f"  Error processing {file_path}: {e}")
    
    # Save output file
    output_path = os.path.join(folder_path, output_file)
    output_wb.save(output_path)
    output_wb.close()
    
    print(f"\nSuccess! Combined {len(excel_files)} files")
    print(f"Total rows (including header): {current_row - 1}")
    print(f"Output saved to: {output_path}")

# Run the script
if __name__ == "__main__":
    folder_path = r"C:\Users\scrib\windev\EtereBridge\output"
    combine_excel_files_convert_formulas(folder_path, 'combined_output.xlsx')